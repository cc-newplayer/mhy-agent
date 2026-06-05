"""
生成人工标注Excel（参照 sentiment_annotation_BV1PnV46DEP4 格式）：
  Sheet1 '人工标注' - 序号/用户/评论内容/人工标注列(留空)
  Sheet2 '模型结果' - 序号/用户/评论内容/点赞数/时间/模型判断/数值得分

标签编码：正面=1, 负面=-1, 中性=0
"""
import json
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT_FILE = "comments_BV1wYdWB5EVF_full_sentiment.json"
OUTPUT_FILE = "sentiment_annotation_BV1wYdWB5EVF.xlsx"
SAMPLE_SIZE = 50

SENTIMENT_SCORE = {"正面": 1, "负面": -1, "中性": 0}


def main():
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        comments = json.load(f)

    # 按情感分组
    groups = {"正面": [], "负面": [], "中性": []}
    for i, c in enumerate(comments):
        s = c.get("sentiment", "中性")
        if s in groups:
            groups[s].append((i, c))

    print(f"数据分布: 正面={len(groups['正面'])}, 负面={len(groups['负面'])}, 中性={len(groups['中性'])}")

    # 各类随机抽取50条
    random.seed(42)
    sampled = []
    for sentiment in ["正面", "负面", "中性"]:
        pool = groups[sentiment]
        n = min(SAMPLE_SIZE, len(pool))
        if n < SAMPLE_SIZE:
            print(f"  警告: {sentiment}类仅有{n}条，不足{SAMPLE_SIZE}条")
        picked = random.sample(pool, n)
        sampled.extend(picked)

    # 打乱顺序
    random.shuffle(sampled)
    print(f"共抽取 {len(sampled)} 条评论")

    # 创建Excel
    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # ========== Sheet1: 人工标注 ==========
    ws1 = wb.active
    ws1.title = "人工标注"

    headers1 = ["序号", "用户", "评论内容", "人工标注（正面/负面/中性）"]
    header_fill1 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill1
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (orig_idx, c) in enumerate(sampled, 2):
        ws1.cell(row=row_idx, column=1, value=row_idx - 1)
        ws1.cell(row=row_idx, column=2, value=c["user"])
        ws1.cell(row=row_idx, column=3, value=c["content"])
        # D列留空，供人工标注
        ws1.cell(row=row_idx, column=4, value=None)

        for col in range(1, 5):
            ws1.cell(row=row_idx, column=col).border = thin_border
            ws1.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 3)
            )

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 60
    ws1.column_dimensions["D"].width = 28
    ws1.freeze_panes = "A2"

    # ========== Sheet2: 模型结果 ==========
    ws2 = wb.create_sheet("模型结果")

    headers2 = ["序号", "用户", "评论内容", "点赞数", "时间", "模型判断", "数值得分"]
    header_fill2 = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill2
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (orig_idx, c) in enumerate(sampled, 2):
        sentiment_text = c.get("sentiment", "中性")
        score = SENTIMENT_SCORE.get(sentiment_text, 0)

        ws2.cell(row=row_idx, column=1, value=row_idx - 1)
        ws2.cell(row=row_idx, column=2, value=c["user"])
        ws2.cell(row=row_idx, column=3, value=c["content"])
        ws2.cell(row=row_idx, column=4, value=c.get("likes", 0))
        ws2.cell(row=row_idx, column=5, value=c.get("time", ""))
        ws2.cell(row=row_idx, column=6, value=sentiment_text)
        ws2.cell(row=row_idx, column=7, value=score)

        for col in range(1, 8):
            ws2.cell(row=row_idx, column=col).border = thin_border
            ws2.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 3)
            )

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 10
    ws2.freeze_panes = "A2"

    # 保存
    wb.save(OUTPUT_FILE)
    print(f"\n已生成: {OUTPUT_FILE}")
    print(f"  Sheet1 '人工标注' - {len(sampled)}条，D列留空待标注")
    print(f"  Sheet2 '模型结果' - 模型判断 + 数值得分")


if __name__ == "__main__":
    main()
