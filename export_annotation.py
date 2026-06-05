"""导出人工标注用的Excel文件"""
import json
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 读取数据
with open("comments_BV13ddcBFEuZ_full_sentiment.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# 按情感分组
positive = [c for c in data if c.get("sentiment") == "正面"]
negative = [c for c in data if c.get("sentiment") == "负面"]
neutral = [c for c in data if c.get("sentiment") == "中性"]

# 每类随机抽50条
random.seed(42)
sample_pos = random.sample(positive, 50)
sample_neg = random.sample(negative, 50)
sample_neu = random.sample(neutral, 50)

# 合并并打乱顺序
all_samples = sample_pos + sample_neg + sample_neu
random.shuffle(all_samples)

# 创建Excel
wb = openpyxl.Workbook()

# === Sheet1: 纯评论（用于人工标注） ===
ws1 = wb.active
ws1.title = "人工标注"
headers1 = ["序号", "用户", "评论内容", "点赞数", "时间", "人工标注（正面/负面/中性）"]
ws1.append(headers1)

# 表头样式
header_font = Font(bold=True)
header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
for col in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill

for i, c in enumerate(all_samples, 1):
    ws1.append([i, c["user"], c["content"], c["likes"], c["time"], ""])

# 列宽
ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 15
ws1.column_dimensions["C"].width = 80
ws1.column_dimensions["D"].width = 8
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 25

# === Sheet2: 评论 + 模型结果（对照用） ===
ws2 = wb.create_sheet("模型结果")
headers2 = ["序号", "用户", "评论内容", "点赞数", "时间", "模型判定", "情感得分"]
ws2.append(headers2)

for col in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill

score_map = {"正面": 1, "中性": 0, "负面": -1}
for i, c in enumerate(all_samples, 1):
    score = score_map.get(c["sentiment"], 0)
    ws2.append([i, c["user"], c["content"], c["likes"], c["time"], c["sentiment"], score])

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 80
ws2.column_dimensions["D"].width = 8
ws2.column_dimensions["E"].width = 20
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 10

# 保存
output_path = "sentiment_annotation_BV1PnV46DEP4.xlsx"
wb.save(output_path)
print(f"已导出: {output_path}")
print(f"共 {len(all_samples)} 条 (正面50 + 负面50 + 中性50)")
