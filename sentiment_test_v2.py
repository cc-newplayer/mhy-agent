"""
使用新提示词对评论进行情感分析（测试用，不覆盖原有结果）
对比新旧结果差异，输出Excel表格
"""
import sys
import io
import json
import time
import os
import concurrent.futures
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import anthropic

API_KEY = os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_AUTH_TOKEN", "")
BASE_URL = os.environ.get("ANTHROPIC_BASE_URL", None)

BATCH_SIZE = 30
MAX_WORKERS = 4
BATCH_INTERVAL = 0.5

INPUT_FILE = "comments_BV1PnV46DEP4_full.json"
OLD_RESULT_FILE = "comments_BV1PnV46DEP4_full_sentiment.json"
NEW_RESULT_FILE = "comments_BV1PnV46DEP4_full_sentiment_v2.json"
DIFF_OUTPUT = "sentiment_diff_BV1PnV46DEP4.xlsx"

NEW_PROMPT_TEMPLATE = """请对以下B站视频评论逐条进行情感分类。
每条评论只需判断为：正面、负面、中性 三选一。

判断标准：

负面（需关注）：
直接批评：骂游戏、骂角色、骂策划（"数值崩了""策划脑子有坑"）。
间接抱怨：表达失望、弃坑、不满（"xx什么时候增强""不想玩了""退坑了""浪费时间"）。
社区冲突信号：角色对比抱怨（"镜流比刃惨""我推待遇不如刃"）——虽然不是在骂游戏，但这类评论是社区情绪分裂的苗头，运营需要知道。
负面推测/谣言："这角色肯定弱""米哈游要凉"。

正面（可忽略或作为参考）：
直接夸奖（"好帅""爱了""必抽"）。
期待、祝福（"期待上线""祝xx生日快乐"——xx是某个游戏角色，如果"生日快乐"是对用户说的，且与游戏无关，应标为中性）。

中性（无需特别关注）：
与游戏无关的社交内容（"前排""第一""哈哈哈"）。
对UP主的个人祝福（"生日快乐""辛苦了"）。
个人计划、理性讨论（不含对角色的拉踩）（"攒了300抽"——若带有明显情绪色彩如攒了这么久终于上线了，可以归为正面、"感觉xx强度在t1"——如果是刃强度不如镜流这种直接表述可归为负面）。
社区维护（"大家别吵了"、"不要引战"）

示例：
中性： 向游侠兄弟致敬；不建议吵架或者要短片在pv和动态底下要
正面：这次pv我反而还有点期待刃的剧情，pv里都加了宗教元素夯爆了；pv质量高
负面（反讽）：太强了，建议削弱（注意识别语气和表情，判断不了归为中性）

评论列表：
{comments_text}

请严格按以下JSON数组格式输出，不要输出其他内容：
[{{"index": 1, "sentiment": "正面"}}, {{"index": 2, "sentiment": "负面"}}, ...]
"""

_progress_lock = threading.Lock()
_completed_count = 0


def analyze_batch(client, comments_batch, batch_id=0):
    comments_text = ""
    for i, comment in enumerate(comments_batch):
        comments_text += f"{i+1}. {comment['content']}\n"

    prompt = NEW_PROMPT_TEMPLATE.format(comments_text=comments_text)

    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.messages.create(
                model="claude-opus-4-7",
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )
            result_text = response.content[0].text.strip()
            start = result_text.find("[")
            end = result_text.rfind("]") + 1
            if start == -1 or end == 0:
                raise ValueError(f"API返回格式异常: {result_text[:200]}")
            return json.loads(result_text[start:end])
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise


def main():
    global _completed_count

    # 读取原始评论
    print(f"读取原始评论: {INPUT_FILE}")
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        comments = json.load(f)
    print(f"  共 {len(comments)} 条")

    # 读取旧结果
    print(f"读取旧结果: {OLD_RESULT_FILE}")
    with open(OLD_RESULT_FILE, "r", encoding="utf-8") as f:
        old_results = json.load(f)

    # 初始化客户端
    client_kwargs = {"api_key": API_KEY}
    if BASE_URL:
        client_kwargs["base_url"] = BASE_URL
    client = anthropic.Anthropic(**client_kwargs)

    # 分批并发处理
    total_batches = (len(comments) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"\n用新提示词跑情感分析 (每{BATCH_SIZE}条一批, {MAX_WORKERS}路并发, 共{total_batches}批)...")

    batches = []
    for batch_idx in range(total_batches):
        s = batch_idx * BATCH_SIZE
        e = min(s + BATCH_SIZE, len(comments))
        batches.append((batch_idx, s, e, comments[s:e]))

    failed_batches = []
    start_time = time.time()

    def process_batch(batch_info):
        global _completed_count
        batch_idx, s, e, batch = batch_info
        try:
            results = analyze_batch(client, batch, batch_idx)
            for item in results:
                idx = item["index"] - 1
                actual_idx = s + idx
                if 0 <= actual_idx < len(comments):
                    sentiment = item["sentiment"]
                    if sentiment not in ("正面", "负面", "中性"):
                        sentiment = "中性"
                    comments[actual_idx]["sentiment_v2"] = sentiment

            with _progress_lock:
                _completed_count += 1
                elapsed = time.time() - start_time
                speed = _completed_count / elapsed * BATCH_SIZE
                print(f"  [{_completed_count}/{total_batches}] 批次{batch_idx+1} 完成 "
                      f"({elapsed:.0f}s, ~{speed:.0f}条/分钟)")
            return True
        except Exception as e:
            with _progress_lock:
                _completed_count += 1
                print(f"  [{_completed_count}/{total_batches}] 批次{batch_idx+1} 失败: {e}")
            failed_batches.append((s, e))
            return False

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = []
        for i, batch_info in enumerate(batches):
            futures.append(executor.submit(process_batch, batch_info))
            if (i + 1) % MAX_WORKERS == 0:
                time.sleep(BATCH_INTERVAL)
        concurrent.futures.wait(futures)

    for s, e in failed_batches:
        for i in range(s, e):
            if "sentiment_v2" not in comments[i]:
                comments[i]["sentiment_v2"] = "中性"

    for c in comments:
        if "sentiment_v2" not in c:
            c["sentiment_v2"] = "中性"

    elapsed_total = time.time() - start_time
    print(f"\n处理完成，总耗时 {elapsed_total:.1f}s ({elapsed_total/60:.1f}分钟)")

    # 保存新结果
    with open(NEW_RESULT_FILE, "w", encoding="utf-8") as f:
        json.dump(comments, f, ensure_ascii=False, indent=2)
    print(f"新结果已保存: {NEW_RESULT_FILE}")

    # 对比差异
    print("\n对比新旧结果...")
    diffs = []
    for i, c in enumerate(comments):
        old_s = old_results[i].get("sentiment", "中性") if i < len(old_results) else "中性"
        new_s = c.get("sentiment_v2", "中性")
        if old_s != new_s:
            diffs.append({
                "content": c["content"],
                "user": c["user"],
                "old": old_s,
                "new": new_s,
            })

    print(f"  总评论: {len(comments)}")
    print(f"  结果不一致: {len(diffs)} 条 ({len(diffs)/len(comments)*100:.1f}%)")

    # 生成Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "情感差异对比"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序号", "用户", "评论内容", "旧提示词结果", "新提示词结果"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 最多输出100条
    show_diffs = diffs[:100]
    for row_idx, d in enumerate(show_diffs, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=d["user"])
        ws.cell(row=row_idx, column=3, value=d["content"])
        ws.cell(row=row_idx, column=4, value=d["old"])
        ws.cell(row=row_idx, column=5, value=d["new"])
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col == 3)
            )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A2"

    # 添加统计摘要sheet
    ws2 = wb.create_sheet("统计摘要")
    ws2.cell(1, 1, "统计项").font = Font(bold=True)
    ws2.cell(1, 2, "数值").font = Font(bold=True)
    ws2.cell(2, 1, "总评论数")
    ws2.cell(2, 2, len(comments))
    ws2.cell(3, 1, "不一致数量")
    ws2.cell(3, 2, len(diffs))
    ws2.cell(4, 1, "不一致比例")
    ws2.cell(4, 2, f"{len(diffs)/len(comments)*100:.1f}%")

    # 变化方向统计
    change_counts = {}
    for d in diffs:
        key = f"{d['old']}→{d['new']}"
        change_counts[key] = change_counts.get(key, 0) + 1

    row = 6
    ws2.cell(row, 1, "变化方向").font = Font(bold=True)
    ws2.cell(row, 2, "数量").font = Font(bold=True)
    for key, count in sorted(change_counts.items(), key=lambda x: x[1], reverse=True):
        row += 1
        ws2.cell(row, 1, key)
        ws2.cell(row, 2, count)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    if len(diffs) > 100:
        note_row = len(show_diffs) + 3
        ws.cell(note_row, 1, f"注：共{len(diffs)}条差异，此处仅展示前100条").font = Font(italic=True, color="666666")

    wb.save(DIFF_OUTPUT)
    print(f"\n差异对比表已生成: {DIFF_OUTPUT}")
    print(f"  Sheet1: 差异评论（最多100条）")
    print(f"  Sheet2: 统计摘要")

    # 打印变化方向
    print("\n变化方向统计:")
    for key, count in sorted(change_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  {key}: {count}条")


if __name__ == "__main__":
    main()
