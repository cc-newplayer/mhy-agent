"""生成混淆矩阵和误判分析文档"""
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook("sentiment_annotation_BV1PnV46DEP4.xlsx")
ws = wb["人工标注"]

# 读取数据
records = []
for row in range(2, ws.max_row + 1):
    idx = ws.cell(row=row, column=1).value
    user = ws.cell(row=row, column=2).value
    content = ws.cell(row=row, column=3).value
    human = ws.cell(row=row, column=4).value
    model = ws.cell(row=row, column=5).value
    agree = ws.cell(row=row, column=6).value

    if human is None or model is None:
        continue
    records.append({
        "idx": idx,
        "user": user,
        "content": content,
        "human": int(human),
        "model": int(model),
        "agree": agree,
    })

print(f"有效标注: {len(records)} 条")

# 标签映射
label_map = {1: "正面", 0: "中性", -1: "负面"}
labels = [1, 0, -1]

# 混淆矩阵
matrix = defaultdict(int)
for r in records:
    matrix[(r["human"], r["model"])] += 1

# 统计
total = len(records)
correct = sum(1 for r in records if r["human"] == r["model"])
accuracy = correct / total * 100

# 每类的precision/recall
report_lines = []
for label in labels:
    tp = matrix[(label, label)]
    fp = sum(matrix[(other, label)] for other in labels if other != label)
    fn = sum(matrix[(label, other)] for other in labels if other != label)
    precision = tp / (tp + fp) * 100 if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
    report_lines.append({
        "label": label_map[label],
        "tp": tp, "fp": fp, "fn": fn,
        "precision": precision, "recall": recall, "f1": f1
    })

# 误判列表
mismatches = [r for r in records if r["human"] != r["model"]]

# 生成 Markdown 报告
md = f"""# 情感分析准确性评估报告

> 视频: BV1PnV46DEP4（崩坏：星穹铁道 千冶·刃 角色PV）
> 样本: {total} 条（正面/中性/负面各50条等比抽样）
> 模型: claude-opus-4-7

## 1. 总体准确率

**{accuracy:.1f}%** ({correct}/{total})

## 2. 混淆矩阵

|  | 模型→正面 | 模型→中性 | 模型→负面 |
|--|-----------|-----------|-----------|
| **人工→正面** | {matrix[(1,1)]} | {matrix[(1,0)]} | {matrix[(1,-1)]} |
| **人工→中性** | {matrix[(0,1)]} | {matrix[(0,0)]} | {matrix[(0,-1)]} |
| **人工→负面** | {matrix[(-1,1)]} | {matrix[(-1,0)]} | {matrix[(-1,-1)]} |

## 3. 各类指标

| 类别 | Precision | Recall | F1 | TP | FP | FN |
|------|-----------|--------|-----|----|----|-----|
"""

for r in report_lines:
    md += f"| {r['label']} | {r['precision']:.1f}% | {r['recall']:.1f}% | {r['f1']:.1f}% | {r['tp']} | {r['fp']} | {r['fn']} |\n"

md += f"""
## 4. 误判分析

共 **{len(mismatches)}** 条误判（{len(mismatches)/total*100:.1f}%）。

### 误判分类统计

"""

# 误判方向统计
error_types = defaultdict(list)
for r in mismatches:
    direction = f"{label_map[r['human']]}→{label_map[r['model']]}"
    error_types[direction].append(r)

for direction, items in sorted(error_types.items(), key=lambda x: -len(x[1])):
    md += f"**{direction}** ({len(items)}条)\n\n"
    for item in items[:5]:  # 每个方向最多列5条
        content = item["content"][:100]
        md += f"- {item['user']}: {content}\n"
    if len(items) > 5:
        md += f"- *...还有{len(items)-5}条*\n"
    md += "\n"

md += """---
*本报告基于人工标注与模型输出对比生成。*
"""

# 保存
output_path = "sentiment_accuracy_report_BV1PnV46DEP4.md"
with open(output_path, "w", encoding="utf-8") as f:
    f.write(md)
print(f"报告已保存: {output_path}")
print(f"准确率: {accuracy:.1f}% ({correct}/{total})")
print(f"误判: {len(mismatches)} 条")
