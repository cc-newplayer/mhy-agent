"""
情感分析准确性评估报告生成器

读取已标注的Excel文件（Sheet1人工标注 + Sheet2模型结果），
生成混淆矩阵、各类指标、误判分析的Markdown报告。

用法:
  python generate_accuracy_report.py sentiment_annotation_BV1wYdWB5EVF.xlsx
"""
import sys
import argparse
import re
from collections import defaultdict
from openpyxl import Workbook, load_workbook


SCORE_TO_LABEL = {1: "正面", 0: "中性", -1: "负面"}
LABEL_TO_SCORE = {"正面": 1, "中性": 0, "负面": -1}
CATEGORIES = ["正面", "中性", "负面"]


def load_data(excel_path):
    """从Excel读取人工标注和模型结果，返回对齐的列表"""
    wb = load_workbook(excel_path)
    ws1 = wb["人工标注"]
    ws2 = wb["模型结果"]

    data = []
    for row in range(2, ws1.max_row + 1):
        idx = ws1.cell(row, 1).value
        if idx is None:
            break

        human_val = ws1.cell(row, 4).value
        if human_val is None:
            continue

        # 支持数值或文字标注
        if isinstance(human_val, (int, float)):
            human_label = SCORE_TO_LABEL.get(int(human_val), "中性")
        else:
            human_label = str(human_val).strip()
            if human_label not in CATEGORIES:
                human_label = "中性"

        user = ws2.cell(row, 2).value or ""
        content = ws2.cell(row, 3).value or ""
        likes = ws2.cell(row, 4).value or 0
        time_str = ws2.cell(row, 5).value or ""

        model_text = ws2.cell(row, 6).value or "中性"
        model_score = ws2.cell(row, 7).value
        if model_score is not None:
            model_label = SCORE_TO_LABEL.get(int(model_score), str(model_text))
        else:
            model_label = str(model_text)

        data.append({
            "idx": idx,
            "user": user,
            "content": content,
            "likes": likes,
            "time": str(time_str),
            "human": human_label,
            "model": model_label,
        })

    return data


def compute_metrics(data):
    """计算混淆矩阵和各类指标"""
    # 混淆矩阵: confusion[human][model] = count
    confusion = defaultdict(lambda: defaultdict(int))
    for d in data:
        confusion[d["human"]][d["model"]] += 1

    total = len(data)
    correct = sum(1 for d in data if d["human"] == d["model"])
    accuracy = correct / total if total > 0 else 0

    # 各类 precision/recall/f1
    metrics = {}
    for cat in CATEGORIES:
        tp = confusion[cat][cat]
        fp = sum(confusion[other][cat] for other in CATEGORIES if other != cat)
        fn = sum(confusion[cat][other] for other in CATEGORIES if other != cat)

        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

        metrics[cat] = {
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "tp": tp,
            "fp": fp,
            "fn": fn,
        }

    return confusion, accuracy, correct, total, metrics


def generate_report(data, video_bv, model_name="claude-opus-4-7"):
    """生成Markdown报告"""
    confusion, accuracy, correct, total, metrics = compute_metrics(data)

    # 误判列表
    errors = defaultdict(list)  # key: "人工→模型"
    for d in data:
        if d["human"] != d["model"]:
            key = f"{d['human']}→{d['model']}"
            errors[key].append(d)

    total_errors = sum(len(v) for v in errors.values())

    # 构建报告
    lines = []
    lines.append("# 情感分析准确性评估报告\n")
    lines.append(f"> 视频: {video_bv}")
    lines.append(f"> 样本: {total} 条（正面/中性/负面各50条等比抽样）")
    lines.append(f"> 模型: {model_name}\n")

    # 1. 总体准确率
    lines.append("## 1. 总体准确率\n")
    lines.append(f"**{accuracy*100:.1f}%** ({correct}/{total})\n")

    # 2. 混淆矩阵
    lines.append("## 2. 混淆矩阵\n")
    lines.append("|  | 模型→正面 | 模型→中性 | 模型→负面 |")
    lines.append("|--|-----------|-----------|-----------|")
    for human_cat in CATEGORIES:
        row_vals = [confusion[human_cat][model_cat] for model_cat in CATEGORIES]
        lines.append(f"| **人工→{human_cat}** | {row_vals[0]} | {row_vals[1]} | {row_vals[2]} |")
    lines.append("")

    # 3. 各类指标
    lines.append("## 3. 各类指标\n")
    lines.append("| 类别 | Precision | Recall | F1 | TP | FP | FN |")
    lines.append("|------|-----------|--------|-----|----|----|-----|")
    for cat in CATEGORIES:
        m = metrics[cat]
        lines.append(
            f"| {cat} | {m['precision']*100:.1f}% | {m['recall']*100:.1f}% | "
            f"{m['f1']*100:.1f}% | {m['tp']} | {m['fp']} | {m['fn']} |"
        )
    lines.append("")

    # 4. 误判分析
    lines.append("## 4. 误判分析\n")
    lines.append(f"共 **{total_errors}** 条误判（{total_errors/total*100:.1f}%）。\n")
    lines.append("### 误判分类统计\n")

    # 按数量排序
    sorted_errors = sorted(errors.items(), key=lambda x: len(x[1]), reverse=True)
    for key, items in sorted_errors:
        lines.append(f"**{key}** ({len(items)}条)\n")
        # 显示前5条
        show_count = min(5, len(items))
        for item in items[:show_count]:
            content_short = item["content"][:50].replace("\n", " ")
            lines.append(f"- {item['user']}: {content_short}")
        if len(items) > show_count:
            lines.append(f"- *...还有{len(items)-show_count}条*")
        lines.append("")

    lines.append("---")
    lines.append("*本报告基于人工标注与模型输出对比生成。*")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="情感分析准确性评估报告生成")
    parser.add_argument("input_file", help="已标注的Excel文件")
    parser.add_argument("--output", help="输出报告文件名")
    parser.add_argument("--model", default="claude-opus-4-7", help="模型名称")
    args = parser.parse_args()

    # 从文件名提取BV号
    bv_match = re.search(r"(BV[\w]+)", args.input_file)
    video_bv = bv_match.group(1) if bv_match else "未知"

    output_file = args.output or f"sentiment_accuracy_report_{video_bv}.md"

    print(f"读取标注文件: {args.input_file}")
    data = load_data(args.input_file)
    print(f"  有效标注: {len(data)} 条")

    if len(data) == 0:
        print("错误: 未找到有效标注数据，请确认Sheet1 D列已填写")
        sys.exit(1)

    report = generate_report(data, video_bv, args.model)

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"\n报告已生成: {output_file}")

    # 打印摘要
    _, accuracy, correct, total, metrics = compute_metrics(data)
    print(f"  总体准确率: {accuracy*100:.1f}% ({correct}/{total})")
    for cat in CATEGORIES:
        m = metrics[cat]
        print(f"  {cat}: P={m['precision']*100:.1f}% R={m['recall']*100:.1f}% F1={m['f1']*100:.1f}%")


if __name__ == "__main__":
    main()
